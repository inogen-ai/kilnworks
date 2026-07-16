import csv
import re
import shutil
import subprocess
from collections.abc import Callable, Iterable, Iterator
from dataclasses import dataclass, field
from pathlib import Path

from bs4 import BeautifulSoup
from docx import Document as DocxDocument
from openpyxl import load_workbook
from PIL import Image
from pypdf import PdfReader

from kilnworks.adapters.media.audio import VIDEO_SUFFIXES, extract_audio
from kilnworks.core.models import Completion
from kilnworks.core.ports import MediaExtractor

TEXT_SUFFIXES = {".md", ".txt"}
PARSED_SUFFIXES = {".pdf", ".docx", ".html", ".htm"}
TABLE_SUFFIXES = {".csv", ".tsv", ".xlsx"}
IMAGE_SUFFIXES = {".png", ".jpg", ".jpeg", ".gif", ".webp"}
MEDIA_SUFFIXES = {".mp3", ".wav", ".m4a", ".mp4", ".mov"}
SUPPORTED_SUFFIXES = (
    TEXT_SUFFIXES | PARSED_SUFFIXES | TABLE_SUFFIXES | IMAGE_SUFFIXES | MEDIA_SUFFIXES
)
MAX_TEXT_CHARS = 10_000_000
# Raw byte ceiling for CSV/TSV files, checked before parsing. The streaming line
# guard (_join_capped) can't help against a single row with no newline: csv.reader
# materializes a whole row before _rows_to_lines ever sees it, so millions of
# delimited fields in one line would balloon memory first. Set well above any table
# that could render under MAX_TEXT_CHARS (CSV text renders on the same order as its
# raw bytes), so legitimate tables are never rejected. XLSX is exempt — its
# cell/column format limits bound a single row, and its compressed size on disk
# doesn't reflect the decompressed content anyway.
MAX_TABLE_BYTES = 4 * MAX_TEXT_CHARS

_IMAGE_MIME_TYPES = {
    ".png": "image/png",
    ".jpg": "image/jpeg",
    ".jpeg": "image/jpeg",
    ".gif": "image/gif",
    ".webp": "image/webp",
}
_MEDIA_MIME_TYPES = {
    ".mp3": "audio/mpeg",
    ".wav": "audio/wav",
    ".m4a": "audio/mp4",
    ".mp4": "video/mp4",
    ".mov": "video/quicktime",
}


class MediaProviderRequired(Exception):
    """Raised by `parse_file` when an image/audio/video file needs a vision or
    transcription provider that isn't configured. Sources catch this the same way
    as any other per-file parse error and surface it as a `SourceFailure`, so one
    unconfigured media file never stops the rest of a batch from ingesting."""

    def __init__(self, suffix: str, kind: str):
        env_var = (
            "KILNWORKS_VISION_PROVIDER" if kind == "vision" else "KILNWORKS_TRANSCRIPTION_PROVIDER"
        )
        super().__init__(
            f"ingesting {suffix} files requires {env_var} to be configured"
        )
        self.suffix = suffix
        self.kind = kind


@dataclass
class ParsedContent:
    """`parse_file`'s return value: the extracted `text`, any Completions spent
    producing it (vision/transcription calls; empty for text/tables/pdf/etc.),
    and best-effort descriptive `metadata` (size/type plus per-format keys)."""

    text: str
    usage: list[Completion] = field(default_factory=list)
    metadata: dict[str, object] = field(default_factory=dict)


# Descriptive MIME types by suffix, spanning every supported format (the
# image/media maps below are reused for extraction and merged in here).
_DOC_MIME_TYPES = {
    ".md": "text/markdown",
    ".txt": "text/plain",
    ".pdf": "application/pdf",
    ".docx": "application/vnd.openxmlformats-officedocument.wordprocessingml.document",
    ".html": "text/html",
    ".htm": "text/html",
    ".csv": "text/csv",
    ".tsv": "text/tab-separated-values",
    ".xlsx": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
}

# Matches the `[MM:SS]` / `[HH:MM:SS]` segment markers the transcription parser
# prefixes onto each block, for counting media segments.
_SEGMENT_RE = re.compile(r"(?m)^\[(\d{1,2}:\d{2}(?::\d{2})?)\]")


def _safe_metadata(fn: Callable[..., dict], *args) -> dict:
    """Run a metadata extractor best-effort: any failure yields no keys rather
    than failing the ingest. Metadata is descriptive, never load-bearing."""
    try:
        return fn(*args) or {}
    except Exception:  # noqa: BLE001 - metadata must never break ingestion
        return {}


def _content_type(suffix: str) -> str | None:
    return _DOC_MIME_TYPES.get(suffix) or _IMAGE_MIME_TYPES.get(suffix) or _MEDIA_MIME_TYPES.get(
        suffix
    )


# Formats whose word count is worth showing; computed post-size-cap in parse_file.
_WORD_COUNT_SUFFIXES = TEXT_SUFFIXES | {".docx", ".html", ".htm"}


def _word_count(text: str) -> dict:
    return {"word_count": len(text.split())}


def _image_metadata(path: Path) -> dict:
    with Image.open(path) as im:
        return {"width": im.width, "height": im.height}


def _table_metadata(path: Path, suffix: str) -> dict:
    if suffix == ".xlsx":
        workbook = load_workbook(path, read_only=True)
        try:
            sheets = workbook.worksheets
            return {
                "sheet_count": len(sheets),
                "row_count": sum((sheet.max_row or 0) for sheet in sheets),
            }
        finally:
            workbook.close()
    with path.open(encoding="utf-8", newline="") as handle:
        return {"row_count": sum(1 for line in handle if line.strip())}


def _duration_seconds(path: Path) -> int | None:
    """Media length in whole seconds via `ffprobe` (bundled with the ffmpeg in the
    Docker image). Returns None when ffprobe is absent or the probe fails — duration
    is descriptive, so a bare non-Docker audio install just omits it."""
    if shutil.which("ffprobe") is None:
        return None
    try:
        completed = subprocess.run(
            ["ffprobe", "-v", "error", "-show_entries", "format=duration",
             "-of", "default=noprint_wrappers=1:nokey=1", str(path)],
            capture_output=True, timeout=30, check=True,
        )
        return round(float(completed.stdout.decode().strip()))
    except Exception:  # noqa: BLE001 - best-effort; absence/failure just omits duration
        return None


def parse_file(path: Path, media: MediaExtractor | None = None) -> ParsedContent:
    suffix = path.suffix.lower()
    usage: list[Completion] = []
    meta: dict[str, object] = {}
    if suffix in TEXT_SUFFIXES:
        text = path.read_text(encoding="utf-8")
    elif suffix == ".pdf":
        # _parse_pdf already enumerates every page; take the count from that same
        # pass rather than re-parsing the whole PDF for it.
        text, page_count = _parse_pdf(path)
        meta["page_count"] = page_count
    elif suffix == ".docx":
        text = _parse_docx(path)
    elif suffix in {".html", ".htm"}:
        text = _parse_html(path)
    elif suffix == ".csv":
        text = _parse_csv(path)
        meta.update(_safe_metadata(_table_metadata, path, suffix))
    elif suffix == ".tsv":
        text = _parse_tsv(path)
        meta.update(_safe_metadata(_table_metadata, path, suffix))
    elif suffix == ".xlsx":
        text = _parse_xlsx(path)
        meta.update(_safe_metadata(_table_metadata, path, suffix))
    elif suffix in IMAGE_SUFFIXES:
        text, completion = _extract_image(path, suffix, media)
        usage.append(completion)
        meta.update(_safe_metadata(_image_metadata, path))
    elif suffix in MEDIA_SUFFIXES:
        text, completion = _extract_media(path, suffix, media)
        usage.append(completion)
        duration = _duration_seconds(path)
        if duration is not None:
            meta["duration_seconds"] = duration
        meta["segment_count"] = len(_SEGMENT_RE.findall(text))
    else:
        raise ValueError(f"unsupported file type: {path.suffix}")

    if len(text) > MAX_TEXT_CHARS:
        msg = (
            f"document too large: extracted {len(text)} chars "
            f"exceeds {MAX_TEXT_CHARS}"
        )
        raise ValueError(msg)

    # Word count is computed here, AFTER the size cap — `str.split` materializes a
    # word list several times the text's size, so running it on unbounded input
    # could OOM a memory-limited worker on a file that would otherwise be cleanly
    # rejected above.
    if suffix in _WORD_COUNT_SUFFIXES:
        meta.update(_word_count(text))

    # Descriptive size/type for every format, best-effort so a metadata hiccup
    # (e.g. the file vanishing between read and stat) never fails an ingest that
    # already produced text — for paid media that would re-bill on retry.
    try:
        meta["size_bytes"] = path.stat().st_size
    except OSError:
        pass
    content_type = _content_type(suffix)
    if content_type:
        meta["content_type"] = content_type
    return ParsedContent(text=text, usage=usage, metadata=meta)


def _check_media_size(path: Path, media: MediaExtractor) -> None:
    size = path.stat().st_size
    if size > media.max_bytes:
        raise ValueError(
            f"media file too large: {path.name} is {size} bytes, exceeding "
            f"KILNWORKS_MAX_MEDIA_BYTES ({media.max_bytes})"
        )


def _extract_image(
    path: Path, suffix: str, media: MediaExtractor | None
) -> tuple[str, Completion]:
    if media is None or media.vision is None:
        raise MediaProviderRequired(suffix, "vision")
    _check_media_size(path, media)
    mime = _IMAGE_MIME_TYPES.get(suffix, "application/octet-stream")
    completion = media.vision.describe(path.read_bytes(), mime, path.name)
    tagged = completion.model_copy(update={"context": "vision"})
    return tagged.text, tagged


def _extract_media(
    path: Path, suffix: str, media: MediaExtractor | None
) -> tuple[str, Completion]:
    if media is None or media.transcription is None:
        raise MediaProviderRequired(suffix, "transcription")
    _check_media_size(path, media)
    raw = path.read_bytes()
    if suffix in VIDEO_SUFFIXES:
        # Whisper-family APIs/models expect audio, not a video container; pull the
        # audio track out with ffmpeg first (extract_audio is a no-op passthrough
        # for suffixes that are already audio). The extracted bytes are WAV, so the
        # name handed to the transcriber must end in .wav too: OpenAI's transcription
        # endpoint validates the multipart filename's EXTENSION against its allowed
        # set (which excludes .mov), and would 400 a valid WAV body carrying a .mov name.
        audio_bytes = extract_audio(raw, suffix)
        mime = "audio/wav"
        transcribe_name = f"{path.stem}.wav"
    else:
        audio_bytes = raw
        mime = _MEDIA_MIME_TYPES.get(suffix, "application/octet-stream")
        transcribe_name = path.name
    completion = media.transcription.transcribe(audio_bytes, mime, transcribe_name)
    tagged = completion.model_copy(update={"context": "transcription"})
    return tagged.text, tagged


def _parse_pdf(path: Path) -> tuple[str, int]:
    # Emit a `[[page:N]]` marker line at the start of each page. The chunker
    # (core/chunking.py) treats each marker as a flush point and tags the following
    # chunk with its 1-based page number, then strips the marker so it never reaches
    # chunk text. See _PAGE_MARKER_RE there. Returns the page count alongside the
    # text so the caller needn't re-parse the PDF just to count pages.
    reader = PdfReader(path)
    pages = reader.pages
    parts = [
        f"[[page:{i + 1}]]\n{(page.extract_text() or '').strip()}"
        for i, page in enumerate(pages)
    ]
    return "\n\n".join(parts).strip(), len(pages)


def _parse_docx(path: Path) -> str:
    document = DocxDocument(str(path))
    return "\n\n".join(p.text for p in document.paragraphs if p.text.strip())


def _parse_html(path: Path) -> str:
    soup = BeautifulSoup(path.read_text(encoding="utf-8"), "html.parser")
    for tag in soup(["script", "style"]):
        tag.decompose()
    return "\n\n".join(soup.stripped_strings)


def _check_table_size(path: Path) -> None:
    size = path.stat().st_size
    if size > MAX_TABLE_BYTES:
        raise ValueError(
            f"document too large: table file is {size} bytes, exceeding {MAX_TABLE_BYTES}"
        )


def _parse_csv(path: Path) -> str:
    _check_table_size(path)
    with path.open(encoding="utf-8", newline="") as handle:
        sample = handle.read(8192)
        handle.seek(0)
        try:
            delimiter = csv.Sniffer().sniff(sample, delimiters=",;\t|").delimiter
        except csv.Error:
            delimiter = ","
        # Stream rows straight into the capped join: a multi-GB CSV is bounded to
        # MAX_TEXT_CHARS of memory and rejected mid-parse, rather than fully
        # materialized via list(csv.reader(...)) before the size check runs.
        return _join_capped(_rows_to_lines(csv.reader(handle, delimiter=delimiter)))


def _parse_tsv(path: Path) -> str:
    _check_table_size(path)
    with path.open(encoding="utf-8", newline="") as handle:
        return _join_capped(_rows_to_lines(csv.reader(handle, delimiter="\t")))


def _parse_xlsx(path: Path) -> str:
    # read_only streams sheet XML row-by-row instead of loading the whole
    # workbook into memory; data_only surfaces cached formula results.
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        return _join_capped(_xlsx_lines(workbook))
    finally:
        workbook.close()


def _xlsx_lines(workbook) -> Iterator[str]:
    """Yield the rendered lines for every non-empty sheet, prefixing each with a
    `# Sheet: <title>` header and a blank separator line between sheets. Streams
    sheet-by-sheet and row-by-row so a huge workbook never fully materializes."""
    first = True
    for sheet in workbook.worksheets:
        rows = (
            [_cell_to_str(value) for value in row]
            for row in sheet.iter_rows(values_only=True)
        )
        lines = _rows_to_lines(rows)
        head = next(lines, None)
        if head is None:
            continue  # no non-empty data rows: omit the sheet (and its header) entirely
        if not first:
            yield ""  # blank line -> "\n\n" separator between sheets
        first = False
        yield f"# Sheet: {sheet.title}"
        yield head
        yield from lines


def _cell_to_str(value: object) -> str:
    return "" if value is None else str(value)


def _rows_to_lines(rows: Iterable[list[str]]) -> Iterator[str]:
    """Yield `label: value | label: value ...` lines from an iterable of rows,
    using the first non-empty row as the header and positional `colN` labels
    wherever a header cell is blank or missing. Blank cells and fully-empty rows
    are skipped. Streams — never materializes all rows — so the caller can bound
    total output size while parsing."""
    header: list[str] | None = None
    for row in rows:
        if not any(cell.strip() for cell in row):
            continue
        if header is None:
            header = row
            continue
        parts = [
            f"{_column_label(header, index)}: {value}"
            for index, cell in enumerate(row)
            if (value := cell.strip())
        ]
        if parts:
            yield " | ".join(parts)


def _join_capped(lines: Iterable[str]) -> str:
    """Join streamed lines with newlines, raising as soon as the accumulated
    length would exceed MAX_TEXT_CHARS. This rejects an oversized table while
    parsing — bounding memory — instead of after building the whole string.
    Reads the module global at call time so tests can lower the cap."""
    out: list[str] = []
    total = 0
    for line in lines:
        total += len(line) + (1 if out else 0)  # +1 for the joining newline
        if total > MAX_TEXT_CHARS:
            raise ValueError(f"document too large: table exceeds {MAX_TEXT_CHARS} chars")
        out.append(line)
    return "\n".join(out)


def _column_label(header: list[str], index: int) -> str:
    if index < len(header) and header[index].strip():
        return header[index].strip()
    return f"col{index + 1}"
